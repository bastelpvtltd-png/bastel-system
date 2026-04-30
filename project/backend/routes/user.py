from flask import Blueprint, request, jsonify, session
import pandas as pd
import os

# Route එක අර්ථ දැක්වීම
users_bp = Blueprint('users', __name__)

def _find_excel():
    base_dir = os.path.dirname(os.path.abspath(__file__))  # routes/
    backend_dir = os.path.abspath(os.path.join(base_dir, '..'))  # backend/
    potential_paths = [
        os.path.abspath(os.path.join(backend_dir, '..', '..', 'data.xlsx')),  # repo root
        os.path.abspath(os.path.join(backend_dir, '..', 'data.xlsx')),        # project/
        os.path.abspath(os.path.join(backend_dir, 'data.xlsx')),              # backend/
        os.path.join(os.getcwd(), 'data.xlsx'),
    ]
    for path in potential_paths:
        if os.path.exists(path):
            return path
    return None

EXCEL_PATH = _find_excel()

@users_bp.route('/api/add-user', methods=['POST'])
def add_user():
    data = request.json
    
    # Frontend එකෙන් එන දත්ත
    full_name = data.get('fullName')
    designation = data.get('designation')
    phone = data.get('phoneNumber')
    email = data.get('email')
    username = data.get('username')
    password = data.get('password')

    try:
        # Excel එක කියවීම
        if EXCEL_PATH and os.path.exists(EXCEL_PATH):
            df = pd.read_excel(EXCEL_PATH, sheet_name='users')
        else:
            return jsonify({"message": "Excel file එක සොයාගත නොහැක!"}), 404

        # Validation
        is_duplicate = df[
            (df['Full Name'] == full_name) | 
            (df['Phone Number'] == phone) | 
            (df['Email'] == email) | 
            (df['Username'] == username)
        ]

        if not is_duplicate.empty:
            return jsonify({"message": "මෙම දත්ත පද්ධතියේ දැනටමත් පවතී!"}), 400

        # අලුත් දත්ත පේළිය
        new_data = {
            "Full Name": full_name,
            "Designation": designation,
            "Phone Number": phone,
            "Email": email,
            "Username": username,
            "Password": password
        }

        # අලුත් පේළිය එකතු කිරීම
        df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)

        # Excel එකට නැවත ලිවීම
        with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='users', index=False)

        return jsonify({"message": "සාර්ථකව දත්ත ඇතුළත් කළා!"}), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"message": "දත්ත ඇතුළත් කිරීමේදී දෝෂයක් සිදුවිය."}), 500

# අලුත් Route එක - මේක කලින් එකෙන් එලියට දාලා තියෙන්නේ
@users_bp.route('/api/get-users', methods=['GET'])
def get_users():
    try:
        if EXCEL_PATH and os.path.exists(EXCEL_PATH):
            df = pd.read_excel(EXCEL_PATH, sheet_name='users')
            # NaN (හිස්) අගයන් JSON වලට යවන්න බැරි නිසා ඒවා හිස් string එකක් කරනවා
            df = df.fillna("") 
            users_list = df.to_dict(orient='records') 
            return jsonify(users_list), 200
        else:
            return jsonify([]), 404
    except Exception as e:
        print(f"Error in get_users: {str(e)}")
        return jsonify({"message": str(e)}), 500
# --- අලුතින් එකතු කරන කොටස --- access eke 

# user.py eke get_access_data function eka mehema wenas karanna
@users_bp.route('/api/get-access-data', methods=['GET'])
def get_access_data():
    try:
        if EXCEL_PATH and os.path.exists(EXCEL_PATH):
            df_users = pd.read_excel(EXCEL_PATH, sheet_name='users')
            user_names = df_users['Full Name'].dropna().unique().tolist()

            df_tabs = pd.read_excel(EXCEL_PATH, sheet_name='TAB')
            tabs_map = {}
            for _, row in df_tabs.iterrows():
                main = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else None
                sub_raw = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else None
                if main:
                    if main not in tabs_map: tabs_map[main] = set()
                    if sub_raw:
                        subs = [s.strip() for s in sub_raw.split(',') if s.strip()]
                        for s in subs: tabs_map[main].add(s)
            
            return jsonify({"users": user_names, "tabs_map": {k: list(v) for k, v in tabs_map.items()}}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500

@users_bp.route('/api/get-user-permissions', methods=['POST'])
def get_user_permissions():
    data = request.json
    full_name = data.get('fullName')
    try:
        if EXCEL_PATH and os.path.exists(EXCEL_PATH):
            df = pd.read_excel(EXCEL_PATH, sheet_name='users')
            user_row = df[df['Full Name'] == full_name]
            if not user_row.empty:
                main_tabs = str(user_row.iloc[0].get('Main', "")).split(',')
                sub_tabs = str(user_row.iloc[0].get('Sub', "")).split(',')
                return jsonify({
                    "selectedMainTabs": [t.strip() for t in main_tabs if t.strip() and t != "nan"],
                    "selectedSubTabs": [t.strip() for t in sub_tabs if t.strip() and t != "nan"]
                }), 200
        return jsonify({"selectedMainTabs": [], "selectedSubTabs": []}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500

@users_bp.route('/api/update-access', methods=['POST'])
def update_access():
    data = request.json
    full_name = data.get('fullName')
    main_tabs_str = ", ".join(data.get('selectedMainTabs', []))
    sub_tabs_str = ", ".join(data.get('selectedSubTabs', []))

    try:
        if EXCEL_PATH and os.path.exists(EXCEL_PATH):
            df = pd.read_excel(EXCEL_PATH, sheet_name='users')
            mask = df['Full Name'] == full_name
            if not df[mask].empty:
                df.loc[mask, 'Main'] = main_tabs_str
                df.loc[mask, 'Sub'] = sub_tabs_str
                with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='w') as writer:
                    df.to_excel(writer, sheet_name='users', index=False)
                return jsonify({"message": "Access සාර්ථකව Update කළා!"}), 200
        return jsonify({"message": "User සොයාගත නොහැක!"}), 404
    except Exception as e:
        return jsonify({"message": str(e)}), 500

# ==========================================
# LOGIN ROUTE - Excel users sheet check
# ==========================================
@users_bp.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username_input = data.get('username', '').strip()
    password_input = data.get('password', '').strip()

    if not username_input or not password_input:
        return jsonify({"success": False, "message": "Username සහ Password අවශ්‍යයි!"}), 400

    try:
        if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
            return jsonify({"success": False, "message": "Excel file eka soyaagatha noheka!"}), 500

        df = pd.read_excel(EXCEL_PATH, sheet_name='users')
        df = df.fillna("")

        # Column E = index 4 = Username, Column F = index 5 = Password
        # Column names වලින් හෝ index වලින් check කරනවා
        if 'Username' in df.columns and 'Password' in df.columns:
            username_col = 'Username'
            password_col = 'Password'
        else:
            # Fallback: column index use කරනවා (E=4, F=5)
            cols = df.columns.tolist()
            username_col = cols[4]
            password_col = cols[5]

        # Normalize columns to string first (Excel may store passwords as numbers e.g. 1234 -> int)
        # This fixes login working in dev but failing after build
        df[username_col] = df[username_col].astype(str).str.strip()
        df[password_col] = df[password_col].astype(str).str.strip()
        # Remove trailing .0 that pandas adds when reading integers from Excel (e.g. "1234.0" -> "1234")
        df[password_col] = df[password_col].str.replace(r'\.0$', '', regex=True)
        df[username_col] = df[username_col].str.replace(r'\.0$', '', regex=True)

        match = df[
            (df[username_col] == username_input) &
            (df[password_col] == password_input)
        ]

        if match.empty:
            return jsonify({"success": False, "message": "Invalid username or password."}), 401

        user_row = match.iloc[0]
        full_name = str(user_row.get('Full Name', '')).strip()
        designation = str(user_row.get('Designation', '')).strip()

        # Admin check - designation "Admin" නම් හෝ username "admin" නම්
        is_admin = (
            username_input.lower() == 'admin' or
            designation.lower() == 'admin'
        )

        # User permissions (main + sub tabs)
        main_tabs_raw = str(user_row.get('Main', '')).strip()
        sub_tabs_raw  = str(user_row.get('Sub', '')).strip()

        main_tabs = [t.strip() for t in main_tabs_raw.split(',') if t.strip() and t != 'nan']
        sub_tabs  = [t.strip() for t in sub_tabs_raw.split(',')  if t.strip() and t != 'nan']

        return jsonify({
            "success": True,
            "username": username_input,
            "fullName": full_name,
            "designation": designation,
            "isAdmin": is_admin,
            "mainTabs": main_tabs,
            "subTabs": sub_tabs,
            "message": "Login successful!"
        }), 200

    except Exception as e:
        print(f"Login Error: {str(e)}")
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500


# ==========================================
# MESSAGES SYSTEM
# ==========================================

MESSAGES_SHEET = 'messages'

def _ensure_messages_sheet():
    """messages sheet නැත්නම් create කරනවා"""
    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        return False
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH)
        if MESSAGES_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(MESSAGES_SHEET)
            ws.append(['id', 'from_user', 'to_user', 'message', 'timestamp', 'is_read'])
            wb.save(EXCEL_PATH)
        return True
    except Exception as e:
        print(f"Messages sheet error: {e}")
        return False

@users_bp.route('/api/messages/send', methods=['POST'])
def send_message():
    data = request.json
    from_user = data.get('from_user', '').strip()
    to_user = data.get('to_user', '').strip()
    message = data.get('message', '').strip()

    if not from_user or not to_user or not message:
        return jsonify({"success": False, "message": "Fields missing"}), 400

    try:
        _ensure_messages_sheet()
        from openpyxl import load_workbook
        from datetime import datetime

        wb = load_workbook(EXCEL_PATH)
        ws = wb[MESSAGES_SHEET]

        # Generate ID
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        new_id = len(rows) + 1

        ws.append([
            new_id,
            from_user,
            to_user,
            message,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            0  # unread
        ])
        wb.save(EXCEL_PATH)
        return jsonify({"success": True, "message": "Sent!"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@users_bp.route('/api/messages/inbox', methods=['POST'])
def get_inbox():
    data = request.json
    username = data.get('username', '').strip()
    try:
        _ensure_messages_sheet()
        df = pd.read_excel(EXCEL_PATH, sheet_name=MESSAGES_SHEET)
        df = df.fillna('')
        inbox = df[df['to_user'].astype(str).str.strip() == username]
        msgs = inbox.sort_values('timestamp', ascending=False).to_dict('records')
        return jsonify({"success": True, "messages": msgs}), 200
    except Exception as e:
        return jsonify({"success": False, "messages": [], "error": str(e)}), 500


@users_bp.route('/api/messages/unread-count', methods=['POST'])
def unread_count():
    data = request.json
    username = data.get('username', '').strip()
    try:
        _ensure_messages_sheet()
        df = pd.read_excel(EXCEL_PATH, sheet_name=MESSAGES_SHEET)
        df = df.fillna('')
        count = len(df[
            (df['to_user'].astype(str).str.strip() == username) &
            (df['is_read'].astype(str).str.strip() == '0')
        ])
        return jsonify({"success": True, "count": count}), 200
    except Exception as e:
        return jsonify({"success": False, "count": 0}), 500


@users_bp.route('/api/messages/mark-read', methods=['POST'])
def mark_read():
    data = request.json
    msg_id = data.get('id')
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH)
        ws = wb[MESSAGES_SHEET]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(msg_id):
                row[5].value = 1
                break
        wb.save(EXCEL_PATH)
        return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False}), 500


@users_bp.route('/api/messages/users-list', methods=['GET'])
def get_users_list():
    """Message කරන්න available users list"""
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name='users')
        df = df.fillna('')
        users = []
        for _, row in df.iterrows():
            uname = str(row.get('Username', '')).strip()
            fname = str(row.get('Full Name', '')).strip()
            if uname:
                users.append({"username": uname, "fullName": fname})
        return jsonify({"success": True, "users": users}), 200
    except Exception as e:
        return jsonify({"success": False, "users": []}), 500


# ==========================================
# PROFILE - GET & UPDATE
# ==========================================

@users_bp.route('/api/profile/get', methods=['POST'])
def get_profile():
    data = request.json
    username = data.get('username', '').strip()
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name='users')
        df = df.fillna('')
        user = df[df['Username'].astype(str).str.strip() == username]
        if user.empty:
            return jsonify({"success": False}), 404
        row = user.iloc[0]
        return jsonify({
            "success": True,
            "fullName": str(row.get('Full Name', '')),
            "designation": str(row.get('Designation', '')),
            "phoneNumber": str(row.get('Phone Number', '')),
            "email": str(row.get('Email', '')),
            "username": str(row.get('Username', '')),
        }), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@users_bp.route('/api/profile/update', methods=['POST'])
def update_profile():
    data = request.json
    username = data.get('username', '').strip()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH)
        ws = wb['users']
        headers = [cell.value for cell in ws[1]]

        col_map = {
            'Full Name': headers.index('Full Name') + 1 if 'Full Name' in headers else None,
            'Designation': headers.index('Designation') + 1 if 'Designation' in headers else None,
            'Phone Number': headers.index('Phone Number') + 1 if 'Phone Number' in headers else None,
            'Email': headers.index('Email') + 1 if 'Email' in headers else None,
        }

        uname_col = headers.index('Username') + 1

        for row in ws.iter_rows(min_row=2):
            if str(row[uname_col - 1].value).strip() == username:
                if col_map['Full Name']: row[col_map['Full Name'] - 1].value = data.get('fullName', '')
                if col_map['Designation']: row[col_map['Designation'] - 1].value = data.get('designation', '')
                if col_map['Phone Number']: row[col_map['Phone Number'] - 1].value = data.get('phoneNumber', '')
                if col_map['Email']: row[col_map['Email'] - 1].value = data.get('email', '')
                break

        wb.save(EXCEL_PATH)
        return jsonify({"success": True, "message": "Profile updated!"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@users_bp.route('/api/profile/change-password', methods=['POST'])
def change_password():
    data = request.json
    username = data.get('username', '').strip()
    current_pw = data.get('currentPassword', '').strip()
    new_pw = data.get('newPassword', '').strip()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH)
        ws = wb['users']
        headers = [cell.value for cell in ws[1]]
        uname_col = headers.index('Username') + 1
        pw_col = headers.index('Password') + 1

        for row in ws.iter_rows(min_row=2):
            if str(row[uname_col - 1].value).strip() == username:
                if str(row[pw_col - 1].value).strip() != current_pw:
                    return jsonify({"success": False, "message": "Current password incorrect!"}), 401
                row[pw_col - 1].value = new_pw
                wb.save(EXCEL_PATH)
                return jsonify({"success": True, "message": "Password changed!"}), 200

        return jsonify({"success": False, "message": "User not found"}), 404
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@users_bp.route('/api/messages/delete', methods=['POST'])
def delete_message():
    data = request.json
    msg_id = data.get('id')
    username = data.get('username', '').strip()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH)
        ws = wb[MESSAGES_SHEET]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(msg_id):
                # Only allow delete if user is sender or receiver
                if str(row[1].value).strip() == username or str(row[2].value).strip() == username:
                    ws.delete_rows(row[0].row)
                    wb.save(EXCEL_PATH)
                    return jsonify({"success": True}), 200
                else:
                    return jsonify({"success": False, "message": "Permission denied"}), 403
        return jsonify({"success": False, "message": "Message not found"}), 404
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@users_bp.route('/api/messages/conversation', methods=['POST'])
def get_conversation():
    """Get all messages between two users (both directions)"""
    data = request.json
    user1 = data.get('user1', '').strip()
    user2 = data.get('user2', '').strip()
    try:
        _ensure_messages_sheet()
        df = pd.read_excel(EXCEL_PATH, sheet_name=MESSAGES_SHEET)
        df = df.fillna('')
        conv = df[
            ((df['from_user'].astype(str).str.strip() == user1) & (df['to_user'].astype(str).str.strip() == user2)) |
            ((df['from_user'].astype(str).str.strip() == user2) & (df['to_user'].astype(str).str.strip() == user1))
        ]
        msgs = conv.sort_values('timestamp').to_dict('records')
        return jsonify({"success": True, "messages": msgs}), 200
    except Exception as e:
        return jsonify({"success": False, "messages": [], "error": str(e)}), 500


@users_bp.route('/api/messages/sent', methods=['POST'])
def get_sent():
    """Get sent messages by a user"""
    data = request.json
    username = data.get('username', '').strip()
    try:
        _ensure_messages_sheet()
        df = pd.read_excel(EXCEL_PATH, sheet_name=MESSAGES_SHEET)
        df = df.fillna('')
        sent = df[df['from_user'].astype(str).str.strip() == username]
        msgs = sent.sort_values('timestamp', ascending=False).to_dict('records')
        return jsonify({"success": True, "messages": msgs}), 200
    except Exception as e:
        return jsonify({"success": False, "messages": []}), 500
